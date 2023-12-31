import tkinter as tk
from tkinter import messagebox
import openpyxl

def is_email_already_registered(email, sheet):
    # for row in sheet.iter_rows(values_only=True):
    for row in sheet.iter_rows(min_row=2):
        if row[0] == email:
            return True
    return False

def register_user():
    email = email_entry.get()
    password = password_entry.get()
    first_name = first_name_entry.get()
    last_name = last_name_entry.get()
    address = address_entry.get()
    phone_number = phone_number_entry.get()

    if email == "" or password == "" or first_name == "" or last_name == "":
        messagebox.showerror("Error", "All fields are required.")
    else:
        excel_file = 'user_data.xlsx'
        try:
            workbook = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            workbook.active.append(['Email', 'Password', 'First Name', 'Last Name', 'Address', 'Phone Number'])

        sheet = workbook.active

        if is_email_already_registered(email, sheet):
            messagebox.showerror("Error", "This email is already registered.")
        else:
            sheet.append([email, password, first_name, last_name, address, phone_number])
            workbook.save(excel_file)
            messagebox.showinfo("Registration Successful", f"User {email} registered successfully!")
            email_entry.delete(0, tk.END)
            password_entry.delete(0, tk.END)
            first_name_entry.delete(0, tk.END)
            last_name_entry.delete(0, tk.END)
            address_entry.delete(0, tk.END)
            phone_number_entry.delete(0, tk.END)
            user_list.insert(tk.END, email)

def get_existing_users():
    excel_file = 'user_data.xlsx'
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        return []

    sheet = workbook.active
    existing_users = []
    # users = [tuple(row) for row in sheet.iter_rows(min_row=2, values_only=True)][1:]
 
    for row in sheet.iter_rows(min_row=2):
        user_data = tuple(cell.value for cell in row)
        existing_users.append(user_data)

    return existing_users

root = tk.Tk()
root.title("User Registration")

label_font = ("Arial", 14)
entry_font = ("Arial", 12)

label_width = 12

first_name_label = tk.Label(root, text="First Name:", font=label_font)
first_name_label.grid(row=0, column=0, padx=10, pady=5)
first_name_entry = tk.Entry(root, font=entry_font)
first_name_entry.grid(row=0, column=1, padx=10, pady=5)

last_name_label = tk.Label(root, text="Last Name:", font=label_font)
last_name_label.grid(row=0, column=2, padx=10, pady=5)
last_name_entry = tk.Entry(root, font=entry_font)
last_name_entry.grid(row=0, column=3, padx=10, pady=5)


email_label = tk.Label(root, text="Email:", font=label_font)
email_label.grid(row=1, column=0, padx=10, pady=5)
email_entry = tk.Entry(root, font=entry_font)
email_entry.grid(row=1, column=1, padx=10, pady=5)

password_label = tk.Label(root, text="Password:", font=label_font)
password_label.grid(row=1, column=2, padx=10, pady=5)
password_entry = tk.Entry(root, show="*", font=entry_font)
password_entry.grid(row=1, column=3, padx=10, pady=5)


address_label = tk.Label(root, text="Address:", font=label_font)
address_label.grid(row=2, column=0, padx=10, pady=5)
address_entry = tk.Entry(root, font=entry_font)
address_entry.grid(row=2, column=1, padx=20, pady=5)

phone_number_label = tk.Label(root, text="Phone Number:", font=label_font)
phone_number_label.grid(row=3, column=0, padx=10, pady=5)
phone_number_entry = tk.Entry(root, font=entry_font)
phone_number_entry.grid(row=3,column=1, padx=10, pady=5)

register_button = tk.Button(root, text="Register", command=register_user, font=label_font)
register_button.grid(row=6, columnspan=6, pady=10)

user_list = tk.Listbox(root, font=label_font)
user_list.grid(row=7, columnspan=8, padx=10, pady=20)

existing_users = get_existing_users()
for user in existing_users:
    user_list.insert(tk.END, user[0])

root.geometry("700x500")
root.mainloop()
